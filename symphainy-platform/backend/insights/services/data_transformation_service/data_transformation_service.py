#!/usr/bin/env python3
"""
Data Transformation Service - Insights Realm

WHAT: Transforms data from source format to target format
HOW: Applies mapping rules and generates output files

Use cases:
- Transform mapped data to Excel format
- Apply data transformations (date formats, type conversions)
- Generate output files with quality flags
"""

import os
import sys
import uuid
import json
from typing import Dict, Any, Optional, List
from datetime import datetime

sys.path.insert(0, os.path.abspath('../../../../../../'))

from bases.realm_service_base import RealmServiceBase

# Try to import pandas and openpyxl for Excel generation
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None


class DataTransformationService(RealmServiceBase):
    """
    Data Transformation Service - Transforms data from source format to target format.
    
    Provides:
    - Apply mapping rules to source data
    - Transform data formats (dates, types, etc.)
    - Generate output files (Excel, JSON, etc.)
    - Include quality flags and citations
    """
    
    def __init__(self, service_name: str, realm_name: str, platform_gateway: Any, di_container: Any):
        """Initialize Data Transformation Service."""
        super().__init__(service_name, realm_name, platform_gateway, di_container)
        
        # Smart City service APIs (will be initialized in initialize())
        self.data_steward = None
        self.file_management = None
        self.nurse = None
    
    async def initialize(self) -> bool:
        """Initialize Data Transformation Service."""
        await super().initialize()
        
        try:
            self.logger.info("🚀 Initializing Data Transformation Service...")
            
            # Get Smart City services
            self.data_steward = await self.get_smart_city_service("DataStewardService")
            self.file_management = await self.get_infrastructure_abstraction("file_management")
            self.nurse = await self.get_smart_city_service("NurseService")
            
            self.logger.info("✅ Data Transformation Service initialized successfully")
            return True
            
        except Exception as e:
            self.logger.error(f"❌ Failed to initialize Data Transformation Service: {e}")
            await self.handle_error_with_audit(e, "initialize")
            return False
    
    async def transform_data(
        self,
        source_data: Dict[str, Any],
        mapping_rules: List[Dict[str, Any]],
        target_schema: Dict[str, Any],
        output_format: str = "excel",
        quality_results: Optional[Dict[str, Any]] = None,
        user_context: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Transform source data to target format using mapping rules.
        
        Args:
            source_data: Source data (records or extracted fields)
            mapping_rules: Mapping rules (source field → target field)
            target_schema: Target schema definition
            output_format: Output format ("excel", "json", "csv")
            quality_results: Optional quality validation results
            user_context: Optional user context
        
        Returns:
        {
            "success": bool,
            "transformed_data": {...},
            "output_file_id": str,  # Generated file
            "transformation_metadata": {
                "fields_mapped": 10,
                "fields_unmapped": 2,
                "confidence_avg": 0.89
            }
        }
        """
        try:
            await self.log_operation_with_telemetry("transform_data_start", success=True)
            
            # Build mapping lookup
            mapping_lookup = {}
            for rule in mapping_rules:
                source_field = rule.get("source_field")
                target_field = rule.get("target_field")
                transformation = rule.get("transformation")
                if source_field and target_field:
                    mapping_lookup[source_field] = {
                        "target_field": target_field,
                        "transformation": transformation,
                        "confidence": rule.get("confidence", 1.0)
                    }
            
            # Get records or extracted fields
            if "records" in source_data:
                # Structured source (list of records)
                records = source_data["records"]
                transformed_records = []
                
                for record in records:
                    transformed_record = {}
                    for source_field, mapping_info in mapping_lookup.items():
                        target_field = mapping_info["target_field"]
                        transformation = mapping_info.get("transformation")
                        
                        if source_field in record:
                            value = record[source_field]
                            
                            # Apply transformation if specified
                            if transformation:
                                value = self._apply_transformation(value, transformation)
                            
                            transformed_record[target_field] = value
                    
                    # Add quality flags if available
                    if quality_results:
                        record_id = record.get("record_id")
                        if record_id:
                            # Find quality info for this record
                            quality_info = next(
                                (r for r in quality_results.get("validation_results", []) if r.get("record_id") == record_id),
                                None
                            )
                            if quality_info:
                                transformed_record["_quality_flag"] = "VALID" if quality_info.get("is_valid") else "INVALID"
                                transformed_record["_quality_score"] = quality_info.get("quality_score", 0.0)
                    
                    transformed_records.append(transformed_record)
                
                transformed_data = {
                    "records": transformed_records,
                    "schema": target_schema
                }
                
            elif "extracted_fields" in source_data:
                # Unstructured source (extracted fields)
                extracted_fields = source_data["extracted_fields"]
                transformed_record = {}
                citations = {}
                confidence_scores = {}
                
                for source_field, mapping_info in mapping_lookup.items():
                    target_field = mapping_info["target_field"]
                    
                    if source_field in extracted_fields:
                        field_data = extracted_fields[source_field]
                        transformed_record[target_field] = field_data.get("value")
                        citations[target_field] = field_data.get("citation")
                        confidence_scores[target_field] = field_data.get("confidence", 0.0)
                
                transformed_data = {
                    "record": transformed_record,
                    "citations": citations,
                    "confidence_scores": confidence_scores,
                    "schema": target_schema
                }
            else:
                return {
                    "success": False,
                    "error": "Unknown source data format"
                }
            
            # Generate output file
            output_file_id = None
            if output_format == "excel":
                output_file_id = await self._generate_excel_file(
                    transformed_data, target_schema, quality_results, citations=transformed_data.get("citations")
                )
            elif output_format == "json":
                output_file_id = await self._generate_json_file(transformed_data)
            elif output_format == "csv":
                output_file_id = await self._generate_csv_file(transformed_data)
            
            # Calculate metadata
            fields_mapped = len(mapping_lookup)
            target_fields = target_schema.get("fields", [])
            fields_unmapped = len(target_fields) - fields_mapped
            
            confidence_scores_list = [rule.get("confidence", 1.0) for rule in mapping_rules]
            confidence_avg = sum(confidence_scores_list) / len(confidence_scores_list) if confidence_scores_list else 0.0
            
            await self.log_operation_with_telemetry("transform_data_complete", success=True, details={
                "fields_mapped": fields_mapped,
                "output_format": output_format
            })
            
            return {
                "success": True,
                "transformed_data": transformed_data,
                "output_file_id": output_file_id,
                "transformation_metadata": {
                    "fields_mapped": fields_mapped,
                    "fields_unmapped": fields_unmapped,
                    "confidence_avg": confidence_avg,
                    "output_format": output_format
                }
            }
            
        except Exception as e:
            self.logger.error(f"❌ Data transformation failed: {e}")
            await self.handle_error_with_audit(e, "transform_data")
            await self.log_operation_with_telemetry("transform_data_complete", success=False, details={"error": str(e)})
            return {
                "success": False,
                "error": str(e)
            }
    
    def _apply_transformation(self, value: Any, transformation: str) -> Any:
        """Apply transformation to value."""
        try:
            if transformation == "date_format":
                # Convert date to YYYY-MM-DD format
                if isinstance(value, str):
                    # Try common date formats
                    from dateutil import parser
                    try:
                        parsed_date = parser.parse(value)
                        return parsed_date.strftime("%Y-%m-%d")
                    except:
                        return value
                return value
            elif transformation == "to_number":
                # Convert to number
                try:
                    if isinstance(value, str):
                        # Remove commas, dollar signs, etc.
                        cleaned = value.replace(",", "").replace("$", "").strip()
                        return float(cleaned) if "." in cleaned else int(cleaned)
                    return float(value) if not isinstance(value, (int, float)) else value
                except:
                    return value
            elif transformation == "to_string":
                return str(value)
            else:
                return value
        except Exception as e:
            self.logger.warning(f"⚠️ Transformation failed: {e}")
            return value
    
    async def _generate_excel_file(
        self,
        transformed_data: Dict[str, Any],
        target_schema: Dict[str, Any],
        quality_results: Optional[Dict[str, Any]],
        citations: Optional[Dict[str, Any]] = None
    ) -> str:
        """Generate Excel file from transformed data."""
        try:
            if not OPENPYXL_AVAILABLE:
                # Fallback to JSON if openpyxl not available
                self.logger.warning("⚠️ openpyxl not available, generating JSON instead")
                return await self._generate_json_file(transformed_data)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Mapped Data"
            
            # Get records
            if "records" in transformed_data:
                records = transformed_data["records"]
            elif "record" in transformed_data:
                records = [transformed_data["record"]]
            else:
                return None
            
            if not records:
                return None
            
            # Get column headers
            headers = list(records[0].keys())
            # Add quality columns if available
            if quality_results:
                if "_quality_flag" not in headers:
                    headers.append("_quality_flag")
                if "_quality_score" not in headers:
                    headers.append("_quality_score")
            if citations:
                if "_citations" not in headers:
                    headers.append("_citations")
            
            # Write headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write data
            for row_idx, record in enumerate(records, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    value = record.get(header)
                    if value is not None:
                        ws.cell(row=row_idx, column=col_idx).value = value
                        
                        # Color code quality flags
                        if header == "_quality_flag" and quality_results:
                            if value == "VALID":
                                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                                    start_color="90EE90", end_color="90EE90", fill_type="solid"
                                )
                            else:
                                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                                    start_color="FFB6C1", end_color="FFB6C1", fill_type="solid"
                                )
            
            # Save to temporary file
            import tempfile
            output_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(output_file.name)
            output_file.close()
            
            # Upload to storage via File Management
            if not self.file_management:
                self.file_management = await self.get_infrastructure_abstraction("file_management")
            
            if self.file_management:
                file_id = f"mapped_data_{uuid.uuid4().hex[:8]}"
                
                # Read file bytes
                with open(output_file.name, 'rb') as f:
                    file_bytes = f.read()
                
                # Upload via File Management
                upload_result = await self.file_management.upload_file(
                    file_data=file_bytes,
                    filename=f"{file_id}.xlsx",
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    metadata={
                        "file_type": "excel",
                        "mapping_output": True,
                        "generated_at": datetime.utcnow().isoformat()
                    }
                )
                
                # Clean up temp file
                import os
                os.unlink(output_file.name)
                
                if upload_result.get("success"):
                    return upload_result.get("file_id") or file_id
            
            return None
            
        except Exception as e:
            self.logger.error(f"❌ Excel file generation failed: {e}")
            return None
    
    async def _generate_json_file(self, transformed_data: Dict[str, Any]) -> str:
        """Generate JSON file from transformed data."""
        try:
            # Convert to JSON
            json_data = json.dumps(transformed_data, indent=2, default=str)
            file_bytes = json_data.encode('utf-8')
            
            # Upload to storage
            if not self.file_management:
                self.file_management = await self.get_infrastructure_abstraction("file_management")
            
            if self.file_management:
                file_id = f"mapped_data_{uuid.uuid4().hex[:8]}"
                
                upload_result = await self.file_management.upload_file(
                    file_data=file_bytes,
                    filename=f"{file_id}.json",
                    content_type="application/json",
                    metadata={
                        "file_type": "json",
                        "mapping_output": True,
                        "generated_at": datetime.utcnow().isoformat()
                    }
                )
                
                if upload_result.get("success"):
                    return upload_result.get("file_id") or file_id
            
            return None
            
        except Exception as e:
            self.logger.error(f"❌ JSON file generation failed: {e}")
            return None
    
    async def _generate_csv_file(self, transformed_data: Dict[str, Any]) -> str:
        """Generate CSV file from transformed data."""
        try:
            if not PANDAS_AVAILABLE:
                self.logger.warning("⚠️ pandas not available, generating JSON instead")
                return await self._generate_json_file(transformed_data)
            
            # Get records
            if "records" in transformed_data:
                records = transformed_data["records"]
            elif "record" in transformed_data:
                records = [transformed_data["record"]]
            else:
                return None
            
            # Convert to DataFrame and CSV
            df = pd.DataFrame(records)
            csv_data = df.to_csv(index=False)
            file_bytes = csv_data.encode('utf-8')
            
            # Upload to storage
            if not self.file_management:
                self.file_management = await self.get_infrastructure_abstraction("file_management")
            
            if self.file_management:
                file_id = f"mapped_data_{uuid.uuid4().hex[:8]}"
                
                upload_result = await self.file_management.upload_file(
                    file_data=file_bytes,
                    filename=f"{file_id}.csv",
                    content_type="text/csv",
                    metadata={
                        "file_type": "csv",
                        "mapping_output": True,
                        "generated_at": datetime.utcnow().isoformat()
                    }
                )
                
                if upload_result.get("success"):
                    return upload_result.get("file_id") or file_id
            
            return None
            
        except Exception as e:
            self.logger.error(f"❌ CSV file generation failed: {e}")
            return None













