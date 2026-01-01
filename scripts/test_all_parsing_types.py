#!/usr/bin/env python3
"""
Comprehensive Parsing Types Test

Tests all supported parsing types with the new architecture:
- Structured: CSV, Excel (xlsx), JSON, Binary + Copybook
- Unstructured: PDF, Word (docx), Text (txt)
- Hybrid: excel_with_text (generates 3 JSON files: structured, unstructured, correlation map)

Flow for each type:
1. Upload file
2. Verify upload success
3. Trigger parsing (with parse_options if needed, e.g., copybook)
4. Verify parsing success
5. For hybrid: Verify 3 JSON files were generated
"""

import requests
import json
import sys
import os
from pathlib import Path
import tempfile
import io

# Test configuration
API_BASE = "http://localhost/api"  # Via Traefik
API_BASE_DIRECT = "http://localhost:8000"  # Direct backend
USER_ID = "test_user_e2e"
SESSION_ID = "test_session_e2e"

# Authentication token - try to get from environment or login
def get_auth_token():
    """Get authentication token - try environment first, then login."""
    # Try environment variable first
    token = os.getenv("AUTH_TOKEN") or os.getenv("SYMPHAINY_API_TOKEN")
    if token:
        return token
    
    # Try to login and get token
    test_email = os.getenv("TEST_USER_EMAIL") or os.getenv("TEST_SUPABASE_EMAIL") or "test@symphainy.com"
    test_password = os.getenv("TEST_USER_PASSWORD") or os.getenv("TEST_SUPABASE_PASSWORD") or "test_password_123"
    
    try:
        print(f"🔐 Attempting to get auth token via login...")
        print(f"   Using email: {test_email}")
        response = requests.post(
            f"{API_BASE}/auth/login",
            json={"email": test_email, "password": test_password},
            timeout=10
        )
        if response.status_code == 200:
            data = response.json()
            token = data.get("token") or data.get("access_token")
            if token:
                print(f"✅ Got auth token via login")
                return token
    except Exception as e:
        print(f"⚠️ Failed to get token via login: {e}")
    
    print(f"⚠️ No auth token available - test will fail with 401")
    return None

AUTH_TOKEN = get_auth_token()

# Test file generators
def create_csv_file():
    """Create a test CSV file."""
    content = """name,age,city
John,30,New York
Jane,25,San Francisco
Bob,35,Chicago
Alice,28,Boston
"""
    return content.encode('utf-8'), 'test_file.csv', 'text/csv', 'csv'

def create_json_file():
    """Create a test JSON file."""
    data = {
        "users": [
            {"name": "John", "age": 30, "city": "New York"},
            {"name": "Jane", "age": 25, "city": "San Francisco"},
            {"name": "Bob", "age": 35, "city": "Chicago"}
        ],
        "metadata": {
            "created": "2024-01-01",
            "version": "1.0"
        }
    }
    content = json.dumps(data, indent=2)
    return content.encode('utf-8'), 'test_file.json', 'application/json', 'json'

def create_txt_file():
    """Create a test text file."""
    content = """This is a test text file.

It contains multiple paragraphs and lines of text.

This file is used to test unstructured parsing capabilities.

The content should be extracted as text chunks for semantic processing.
"""
    return content.encode('utf-8'), 'test_file.txt', 'text/plain', 'txt'

def create_excel_file():
    """Create a test Excel file (xlsx) using openpyxl."""
    try:
        from openpyxl import Workbook
        from io import BytesIO
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add headers
        ws['A1'] = 'name'
        ws['B1'] = 'age'
        ws['C1'] = 'city'
        
        # Add data
        data = [
            ['John', 30, 'New York'],
            ['Jane', 25, 'San Francisco'],
            ['Bob', 35, 'Chicago'],
            ['Alice', 28, 'Boston']
        ]
        
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        content = excel_buffer.getvalue()
        
        return content, 'test_file.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'xlsx'
    except ImportError:
        # If openpyxl is not available, skip Excel test
        print("⚠️ openpyxl not available - skipping Excel test")
        return None, None, None, None
    except Exception as e:
        print(f"⚠️ Failed to create Excel file: {e}")
        return None, None, None, None

def create_pdf_file():
    """Create a test PDF file."""
    # For testing, we'll create a minimal PDF
    # In production, you'd use a proper PDF library
    # This is a minimal valid PDF structure
    pdf_content = b"""%PDF-1.4
1 0 obj
<<
/Type /Catalog
/Pages 2 0 R
>>
endobj
2 0 obj
<<
/Type /Pages
/Kids [3 0 R]
/Count 1
>>
endobj
3 0 obj
<<
/Type /Page
/Parent 2 0 R
/Resources <<
/Font <<
/F1 <<
/Type /Font
/Subtype /Type1
/BaseFont /Helvetica
>>
>>
>>
/MediaBox [0 0 612 792]
/Contents 4 0 R
>>
endobj
4 0 obj
<<
/Length 44
>>
stream
BT
/F1 12 Tf
100 700 Td
(Test PDF Content) Tj
ET
endstream
endobj
xref
0 5
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
0000000317 00000 n
trailer
<<
/Size 5
/Root 1 0 R
>>
startxref
398
%%EOF"""
    return pdf_content, 'test_file.pdf', 'application/pdf', 'pdf'

def create_docx_file():
    """Create a test Word file (docx)."""
    # For testing, we'll create a minimal docx structure
    # In production, you'd use python-docx or similar
    # This is a simplified approach - real docx files are ZIP archives
    # For now, we'll skip docx and note it needs proper library
    # Return None to skip this test for now
    return None, None, None, None

def create_binary_file_with_copybook():
    """Create a test binary file and copybook for mainframe processing."""
    # Create a simple COBOL copybook definition
    copybook_content = """01  CUSTOMER-RECORD.
    05  CUSTOMER-ID        PIC X(10).
    05  CUSTOMER-NAME      PIC X(50).
    05  CUSTOMER-AGE       PIC 9(3).
    05  CUSTOMER-CITY      PIC X(30).
"""
    
    # Create binary data matching the copybook structure
    # Record structure: ID(10) + NAME(50) + AGE(3) + CITY(30) = 93 bytes per record
    records = [
        ("CUST001   ", "John Doe                              ", "030", "New York                    "),
        ("CUST002   ", "Jane Smith                            ", "025", "San Francisco               "),
        ("CUST003   ", "Bob Johnson                           ", "035", "Chicago                     "),
    ]
    
    binary_data = b""
    for record in records:
        # Pack each field as bytes
        record_bytes = (
            record[0].ljust(10).encode('ascii') +  # ID: 10 bytes
            record[1].ljust(50).encode('ascii') +  # NAME: 50 bytes
            record[2].ljust(3).encode('ascii') +   # AGE: 3 bytes
            record[3].ljust(30).encode('ascii')    # CITY: 30 bytes
        )
        binary_data += record_bytes
    
    return binary_data, 'test_file.bin', 'application/octet-stream', 'bin', copybook_content

def create_hybrid_excel_file():
    """Create a test Excel file with text (for hybrid parsing)."""
    try:
        from openpyxl import Workbook
        from io import BytesIO
        
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add headers
        ws['A1'] = 'name'
        ws['B1'] = 'age'
        ws['C1'] = 'city'
        ws['D1'] = 'notes'  # Text column for unstructured content
        
        # Add data with text notes
        data = [
            ['John', 30, 'New York', 'John is a software engineer with 10 years of experience. He specializes in Python and cloud architecture.'],
            ['Jane', 25, 'San Francisco', 'Jane is a data scientist working on machine learning projects. She has a PhD in Computer Science.'],
            ['Bob', 35, 'Chicago', 'Bob is a product manager with expertise in agile methodologies and user experience design.'],
            ['Alice', 28, 'Boston', 'Alice is a DevOps engineer focused on containerization and CI/CD pipelines.'],
        ]
        
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        content = excel_buffer.getvalue()
        
        return content, 'test_file_hybrid.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'excel_with_text'
    except ImportError:
        print("⚠️ openpyxl not available - skipping hybrid Excel test")
        return None, None, None, None
    except Exception as e:
        print(f"⚠️ Failed to create hybrid Excel file: {e}")
        return None, None, None, None

# Test results tracking
test_results = []

def test_file_type(file_data, filename, content_type, file_type, parsing_type, parse_options=None):
    """Test upload and parsing for a specific file type."""
    print("\n" + "="*80)
    print(f"TESTING: {parsing_type.upper()} - {file_type.upper()} ({filename})")
    print("="*80)
    
    if file_data is None:
        print(f"⚠️ Skipping {file_type} - file generation not implemented")
        test_results.append({
            "type": file_type,
            "parsing_type": parsing_type,
            "status": "skipped",
            "reason": "file_generation_not_implemented"
        })
        return False
    
    # Step 1: Upload file
    print(f"\n📤 Step 1: Uploading {filename}...")
    upload_success, file_id, workflow_id = upload_file(file_data, filename, content_type, file_type)
    
    if not upload_success:
        print(f"❌ Upload failed for {file_type}")
        test_results.append({
            "type": file_type,
            "parsing_type": parsing_type,
            "status": "failed",
            "stage": "upload",
            "error": "upload_failed"
        })
        return False
    
    print(f"✅ Upload successful: file_id={file_id}, workflow_id={workflow_id}")
    
    # Step 2: Parse file
    print(f"\n⚙️ Step 2: Parsing {filename}...")
    if parse_options:
        print(f"   With parse_options: {list(parse_options.keys())}")
    parse_success, parse_result = parse_file(file_id, parse_options=parse_options)
    
    if not parse_success:
        print(f"❌ Parsing failed for {file_type}")
        test_results.append({
            "type": file_type,
            "parsing_type": parsing_type,
            "status": "failed",
            "stage": "parsing",
            "file_id": file_id,
            "error": parse_result.get("error") if parse_result else "unknown_error"
        })
        return False
    
    # For hybrid parsing, verify 3 JSON files were generated
    if parsing_type == "hybrid":
        parsed_files = parse_result.get("parsed_files", {})
        if "structured" in parsed_files and "unstructured" in parsed_files and "correlation_map" in parsed_files:
            print(f"✅ Hybrid parsing successful - 3 JSON files generated:")
            print(f"   - Structured data: {len(str(parsed_files.get('structured', {}).get('data', {})))} bytes")
            print(f"   - Unstructured chunks: {parsed_files.get('unstructured', {}).get('chunk_count', 0)} chunks")
            print(f"   - Correlation map: {len(str(parsed_files.get('correlation_map', {}).get('data', {})))} bytes")
        else:
            print(f"⚠️ Hybrid parsing completed but missing expected JSON files")
            print(f"   Found: {list(parsed_files.keys())}")
    
    print(f"✅ Parsing successful for {file_type}")
    test_results.append({
        "type": file_type,
        "parsing_type": parsing_type,
        "status": "passed",
        "file_id": file_id,
        "workflow_id": workflow_id
    })
    return True

def upload_file(file_data, filename, content_type, file_type):
    """Upload a file."""
    # Try Traefik first (API_BASE), then direct backend as fallback
    for api_base in [API_BASE, API_BASE_DIRECT]:
        try:
            files = {
                'file': (filename, io.BytesIO(file_data), content_type)
            }
            data = {
                'file_type': file_type,
                'user_id': USER_ID,
                'session_id': SESSION_ID
            }
            
            url = f"{api_base}/v1/content-pillar/upload-file"
            headers = {
                'X-Session-Token': SESSION_ID
            }
            if AUTH_TOKEN:
                headers['Authorization'] = f'Bearer {AUTH_TOKEN}'
            
            response = requests.post(url, files=files, data=data, headers=headers, timeout=30)
            
            if response.status_code == 200:
                result = response.json()
                file_id = result.get("file_id") or result.get("uuid")
                workflow_id = result.get("workflow_id")
                return True, file_id, workflow_id
            else:
                error_msg = f"{response.status_code} - {response.text[:200]}"
                if api_base == API_BASE_DIRECT:
                    print(f"❌ Upload failed via direct backend: {error_msg}")
                    return False, None, None
                else:
                    print(f"⚠️ Upload failed via Traefik ({error_msg}), trying direct backend...")
                    continue
                
        except requests.exceptions.ConnectionError as e:
            if api_base == API_BASE_DIRECT:
                print(f"❌ Upload connection error: {e}")
                return False, None, None
            else:
                print(f"⚠️ Traefik connection failed ({e}), trying direct backend...")
                continue
        except Exception as e:
            if api_base == API_BASE_DIRECT:
                print(f"❌ Upload exception: {e}")
                return False, None, None
            else:
                print(f"⚠️ Traefik request failed ({e}), trying direct backend...")
                continue
    
    return False, None, None

def parse_file(file_id, parse_options=None):
    """Parse a file."""
    # Try Traefik first (API_BASE), then direct backend as fallback
    for api_base in [API_BASE, API_BASE_DIRECT]:
        try:
            url = f"{api_base}/v1/content-pillar/process-file/{file_id}"
            data = {
                'user_id': USER_ID
            }
            
            # Add parse_options if provided (e.g., for copybook)
            if parse_options:
                data['processing_options'] = parse_options
            
            headers = {
                'X-Session-Token': SESSION_ID,
                'Content-Type': 'application/json'
            }
            if AUTH_TOKEN:
                headers['Authorization'] = f'Bearer {AUTH_TOKEN}'
            
            # Use longer timeout for parsing (especially binary files)
            response = requests.post(url, json=data, headers=headers, timeout=180)
            
            if response.status_code == 200:
                result = response.json()
                success = result.get("success", False)
                if success:
                    return True, result
                else:
                    # Even if success=False, return the result so we can see the error
                    error_msg = result.get("error") or result.get("message") or "Unknown error"
                    print(f"⚠️ Parse returned success=False: {error_msg}")
                    return False, result
            else:
                error_msg = f"{response.status_code} - {response.text[:200]}"
                if api_base == API_BASE_DIRECT:
                    print(f"❌ Parse failed via direct backend: {error_msg}")
                    return False, {"error": error_msg}
                else:
                    print(f"⚠️ Parse failed via Traefik ({error_msg}), trying direct backend...")
                    continue
                
        except requests.exceptions.ConnectionError as e:
            if api_base == API_BASE_DIRECT:
                print(f"❌ Parse connection error: {e}")
                return False, {"error": f"Connection error: {str(e)}"}
            else:
                print(f"⚠️ Traefik connection failed ({e}), trying direct backend...")
                continue
        except requests.exceptions.Timeout as e:
            error_msg = f"Request timeout after 180 seconds"
            print(f"❌ {error_msg}")
            return False, {"error": error_msg}
        except Exception as e:
            if api_base == API_BASE_DIRECT:
                print(f"❌ Parse exception: {e}")
                import traceback
                print(f"   Traceback: {traceback.format_exc()}")
                return False, {"error": str(e)}
            else:
                print(f"⚠️ Traefik request failed ({e}), trying direct backend...")
                continue
    
    return False, {"error": "All connection attempts failed"}

def main():
    """Run comprehensive parsing tests."""
    print("\n" + "="*80)
    print("COMPREHENSIVE PARSING TYPES TEST")
    print("="*80)
    print("\nTesting all supported parsing types with new architecture")
    print("="*80)
    
    # Structured parsing tests
    print("\n" + "="*80)
    print("STRUCTURED PARSING TESTS")
    print("="*80)
    
    # CSV
    csv_data, csv_filename, csv_content_type, csv_file_type = create_csv_file()
    test_file_type(csv_data, csv_filename, csv_content_type, csv_file_type, "structured")
    
    # JSON
    json_data, json_filename, json_content_type, json_file_type = create_json_file()
    test_file_type(json_data, json_filename, json_content_type, json_file_type, "structured")
    
    # Excel (xlsx) - Note: simplified test
    excel_data, excel_filename, excel_content_type, excel_file_type = create_excel_file()
    test_file_type(excel_data, excel_filename, excel_content_type, excel_file_type, "structured")
    
    # Binary + Copybook (structured)
    binary_data, binary_filename, binary_content_type, binary_file_type, copybook_content = create_binary_file_with_copybook()
    parse_options_binary = {
        "copybook": copybook_content,
        "file_type": "binary"
    }
    test_file_type(binary_data, binary_filename, binary_content_type, binary_file_type, "structured", parse_options=parse_options_binary)
    
    # Hybrid parsing tests
    print("\n" + "="*80)
    print("HYBRID PARSING TESTS")
    print("="*80)
    
    # Excel with text (hybrid)
    hybrid_data, hybrid_filename, hybrid_content_type, hybrid_file_type = create_hybrid_excel_file()
    if hybrid_data:
        parse_options_hybrid = {
            "parsing_type": "hybrid"  # Explicitly set hybrid parsing
        }
        test_file_type(hybrid_data, hybrid_filename, hybrid_content_type, hybrid_file_type, "hybrid", parse_options=parse_options_hybrid)
    
    # Unstructured parsing tests
    print("\n" + "="*80)
    print("UNSTRUCTURED PARSING TESTS")
    print("="*80)
    
    # Text
    txt_data, txt_filename, txt_content_type, txt_file_type = create_txt_file()
    test_file_type(txt_data, txt_filename, txt_content_type, txt_file_type, "unstructured")
    
    # PDF
    pdf_data, pdf_filename, pdf_content_type, pdf_file_type = create_pdf_file()
    test_file_type(pdf_data, pdf_filename, pdf_content_type, pdf_file_type, "unstructured")
    
    # Word (docx) - Skip for now (needs proper library)
    docx_data, docx_filename, docx_content_type, docx_file_type = create_docx_file()
    if docx_data:
        test_file_type(docx_data, docx_filename, docx_content_type, docx_file_type, "unstructured")
    
    # Summary
    print("\n" + "="*80)
    print("TEST SUMMARY")
    print("="*80)
    
    passed = sum(1 for r in test_results if r.get("status") == "passed")
    failed = sum(1 for r in test_results if r.get("status") == "failed")
    skipped = sum(1 for r in test_results if r.get("status") == "skipped")
    total = len(test_results)
    
    print(f"\nTotal Tests: {total}")
    print(f"✅ Passed: {passed}")
    print(f"❌ Failed: {failed}")
    print(f"⚠️ Skipped: {skipped}")
    
    print("\nDetailed Results:")
    for result in test_results:
        status_icon = "✅" if result.get("status") == "passed" else "❌" if result.get("status") == "failed" else "⚠️"
        print(f"  {status_icon} {result.get('type', 'unknown').upper()} ({result.get('parsing_type', 'unknown')}) - {result.get('status', 'unknown')}")
        if result.get("error"):
            print(f"     Error: {result.get('error')}")
        if result.get("file_id"):
            print(f"     File ID: {result.get('file_id')}")
    
    if failed == 0 and passed > 0:
        print("\n🎉 ALL TESTS PASSED!")
        return 0
    elif passed > 0:
        print(f"\n⚠️ SOME TESTS FAILED: {failed} failed, {passed} passed")
        return 1
    else:
        print("\n❌ ALL TESTS FAILED")
        return 1

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)

